# import modules
import arcpy, os, re, csv
from arcpy import env
from arcpy.sa import *
from dbfpy import dbf
import openpyxl


# Who: This script is the combined work of Elizabeth Margolin, Benjamin Stewart, and Keith Garrett at the WBG

# Why: To quickly obtain raw zonal stats on postprocessed DMSP-OLS "Nightlights" data (
# Annual Ave and RadCal) to get to usable table in Excel or STATA, in an automated way.

# What: It uses ArcGIS to take a table with x,y coordinates, transforms it into a feature class,
# Then creates a buffer unique in size to each point on the basis of a user specified field.
# It calculates zonal stats on each of those points against all the rasters in a specified folder,
# then assembles those results into a single table

arcpy.ImportToolbox('/Desktop/NTL_TEST/SpatialAnalystSupplementalTools/Spatial Analyst Supplemental Tools.pyt')

def main():
    # Check out the ArcGIS Spatial Analyst extension license
    arcpy.CheckOutExtension("Spatial")

    #The below are variables that the USER will have to set, based on their preferred file path
    w_path = "/Desktop/NTL_TEST/" #output folder
    #eval_sites should be formatted with columns titled: id, latitude, longitude, size
    eval_sites = "/Desktop/NTL_TEST/GPS_final.csv" #sites to be evaluated
    #In eval_sites x and y coords should be saved as longitude and latitude, if you want to change this
    #change the 2nd and 3rd variable in line 48
    fcEval_sites = "/Desktop/NTL_TEST/Eval_Sites.shp" #Shape file that will contain sites
    fcEval_sites_proj = "/Desktop/NTL_TEST/Eval_Sites_proj.shp"
    rawNTLFolder = "/Desktop/NTL_TEST/RawNTLCorrected/"  # Folder that contains the input Nighttime lights images
    radcalNTLFolder = "/Desktop/NTL_TEST/RadCal_Corrected/" #Folder that contain the rad_cal NTL
    elEval_sites = "EVAL_Layer"  # Temp Layer Name
    num_inc = 1 #number of buffers
    inc_size = 10000 #Radius increase for each buffer, given in meters
    outxls_file = ""
    pref_stats= ['SUM','MEAN','STD'] #SUM, MEAN, STD are all uniquely options; ALL IS NOT AN OPTION
    arcpy.env.workspace = r'C:\2_Data\Data_Raw.gdb'

    ntlFiles = getNTLFiles(rawNTLFolder, radcalNTLFolder)

    # If the projected version doesn't already exist
    if not arcpy.Exists(fcEval_sites_proj):
        # Convert lat/long to table to points
        sr = arcpy.SpatialReference(4326)  # GCS WGS84
        arcpy.MakeXYEventLayer_management(eval_sites, "longitude", "latitude", elEval_sites, sr, "")
        arcpy.MakeFeatureLayer_management(elEval_sites, "TEMP", "", "", "")
        arcpy.CopyFeatures_management("TEMP",fcEval_sites)
        
        # create a spatial reference object for the output coordinate system, in this case Albers Equal Area Conic
        out_CS = arcpy.SpatialReference(102022)
        arcpy.Project_management(fcEval_sites, fcEval_sites_proj, out_CS)

    # Loop through and create a shape_file for each buffer layer
    ntlVals = []
    for buffer_num in range(1, num_inc+1):
        out_Buff = os.path.join(w_path, 'SCT_Buffer_%s.shp' % (buffer_num))
        if not os.path.exists(out_Buff):
            arcpy.Buffer_analysis(fcEval_sites_proj, out_Buff, inc_size*buffer_num)
        #Have NTLVals be a list of the dbf files
        for file in summarizeNTLData(out_Buff, "id", ntlFiles, w_path):
            ntlVals.append(file)

    #Just merge all the NTLVals files into one dbf and merege
    summarizeDbf(ntlVals,'id',pref_stats, outxls_file)


'''Open all the dbf files in dbf list and extract all relevant values
---dbfList: list of dbf files to process, these should be from zonal statistics in most cases
---idField: Field that identifies the individual zonal features
---sumVals: List of equal length to dbf list describing the fields to extract from dbf. Use -1 to get just the sums
--RETURNS: An excel spreadsheet with the specificed stats run on each NTL file
'''
def summarizeDbf(dbfList, idField, stats_choice, output_file):
    wb = openpyxl.Workbook() # establishes the workbook, and sets the base as main sheet
    main_sheet = wb.active

    main_sheet['A1'] = "ID"
    col_num = 2
    ids = {}
    # set up the excel file
    for dbf_inlist in dbfList:
        dbfFile = dbf.Dbf(dbf_inlist)
        
        for rec in dbfFile:
            if rec['id'] not in ids:
                main_sheet.cell(None, len(ids) + 2, 1).value = rec['id']
                ids[rec['id']] = len(ids) + 2
                
            name = os.path.basename(dbf_inlist)
            
            for i in range(0,stats_choice):
                main_sheet.cell(None, 1, col_num).value = name + stats_choice[i]
                row = ids[rec['id']]
                main_sheet.cell(None,row,col_num).value = rec[stats_choice[i]]
                col_num += len(stats_choice)

    wb.save(output_file)

    for f in dbfList:
        arcpy.Delete_management(f)

# Run zonal statistics on all nighttime lights files (in the inFiles list)
###inShp - input shape in which to summarize nighttime lights data
###idField - unique identifier for the inShp
###inFiles - list of nighttime lights files (generated from getNTLFiles()
###stats - preferred stats
###base_path - essentially w_path
###RETURNS - a list of dbf files that will then be processed by summarize dbf
def summarizeNTLData(inShp, idField, inFiles, base_path):
    out_files = []
    tempVals = []
    arcpy.env.overwriteOutput = True #DO NOT TOUCH THIS!
    
    for ras_file in inFiles:
        tempVals.append(ras_file)
        fName = os.path.basename(ras_file)
        inR = arcpy.Raster(ras_file)
        outTable = base_path + fName[:-4].replace(".","") + ("%s.dbf" % inShp[-5])
        
        if not os.path.exists(outTable):
            arcpy.ZonalStatisticsAsTable02_sas(inShp, idField, inR, outTable)
        out_files.append(outTable)
        print("Finshed Processing " + fName[:-4] + ("%s" % inShp[-5]))

    return (out_files)


# Generate a list of nighttime lights files from the folders given up in main
def getNTLFiles(rawFolder,radCalFolder):
    ntlFiles = []

    for f in os.listdir(rawFolder):
        if re.search("ElvidgeCorrected_int.tif$", f):
            ntlFiles.append(os.path.join(rawFolder,f))

    for f in os.listdir(radCalFolder):
        if re.search("_vis.tif$", f):
            ntlFiles.append(os.path.join(radCalFolder, f))

    return (ntlFiles)


if __name__ == '__main__':
    main()
