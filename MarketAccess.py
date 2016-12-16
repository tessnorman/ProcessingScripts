from dbfread import DBF
from openpyxl import load_workbook
import math
import csv

# Script by Charles Fox
# Assesses the Market Potential of a given Adminstrative unit on the basis of the population weighted euclidian distance of other admin units
# Based on a formula worked out by Therese Norman and Charles Fox, based on (Dichemann, Fay 2004)
###### INPUTS #########

#work path: "C:\Users\wb493355\Documents\Excel\Bangaldesh"
path = "C:\Users\charl\Documents\Market Access"
inputdbf = "\BGDAdmin3_ProxTable.dbf"
inputpop = "\Upazila populations.xlsx"
operdbf = path + inputdbf
operpop_wb = path + inputpop
operpop_sheet = 'Output'
rs_row = 2
re_row = 529
test_tuples = 3
practice_mode = 'no'
practice_record_total = 5000
lambder_list = [1000,100,50,25,10,5,2,1,0.5,0.25]

###### STILL TO DO #########
"""
- organise as dicts instead of lists
- front end: download shapefile, run dist matrix on shapefile
- remove .dbf references when distmatrix as list object
- make 'fuzzy match' file checker for pop names
- back end: merge csv onto shapefile
- make dist matrix run on roads not euclidian
"""
###### DISTANCE MATRIX #########
distmatrix = DBF(operdbf, load=True)

if practice_mode == 'yes':
    record_total = practice_record_total
else:
    record_total = len(distmatrix)

market_dist_tuples_list = []

for z in range(0,record_total):
	input_district = distmatrix.records[z]['IN_FID']
	output_district = distmatrix.records[z]['NEAR_FID']
	distance = distmatrix.records[z]['NEAR_DIST']
	zz = (input_district,output_district,distance)
	market_dist_tuples_list.append(zz)

print "\nfirst few tuples of market_dist_tuples:"
print market_dist_tuples_list[0:test_tuples]

###### POPULATION #########

pop_wb = load_workbook(operpop_wb)
pop_wb_ws = pop_wb.get_sheet_by_name(operpop_sheet)
population_tuples_list = []
for rows in range(rs_row,re_row):
    district_ID = int(pop_wb_ws.cell(row=rows,column=1).value)
    district_name = str(pop_wb_ws.cell(row=rows,column=2).value)
    district_pop = int(pop_wb_ws.cell(row=rows,column=3).value)
    pop_tuple = [district_ID,district_name,district_pop]
    population_tuples_list.append(pop_tuple)

print "\nfirst few tuples of population_tuples_list:"
print population_tuples_list[0:test_tuples]

###### JOIN DIST TO POP #########

joined_tuples = []

for dist_tuple in market_dist_tuples_list:
    for pop_tuple in population_tuples_list:
        if dist_tuple[1] == pop_tuple[0]:
            new_list = [dist_tuple[0],dist_tuple[1],dist_tuple[2],pop_tuple[0],pop_tuple[1],pop_tuple[2]]
            joined_tuples.append(new_list)
        else:
            pass
print "\nfirst few tuples of joined_tuples_list:"
print joined_tuples[0:test_tuples]
print "\ntotal number of joined tuples is %d" % len(joined_tuples)

###### ADD DECAYED POPULATION VALUE TO JOINED TUPLES #########

def decayed_distance_calculator(x, lambder):
    a = joined_tuples[x][2]
    b = (-1 * float(lambder) * a)
    c = joined_tuples[x][5] * math.exp(b)
    joined_tuples[x].append(c)
    return joined_tuples[x]

for new_tupless in range(0,record_total):
    for lambder in lambder_list:
        decayed_distance_calculator(new_tupless, lambder)

print "\nfirst few tuples of joined_tuples:"
print joined_tuples[0:test_tuples]

###### CALCULATE MARKET ACCESS FOR EACH DISTRICT #########
#start with list from the population file as it includes
#most of the key things (ID, name, pop)
# By Charles Fox(y)
def aggregator(g,lambder_no):
    sumlist = []
    for a in range(0,record_total):
        if joined_tuples[a][0] == population_tuples_list[g][0]:
            sumlist.append(joined_tuples[a][lambder_no])
        else:
            pass
    b = sum(sumlist) + population_tuples_list[g][2]
    population_tuples_list[g].append(b)

districts = len(population_tuples_list)

lambdas = []
for x in range(0,len(lambder_list)):
    p = x + 6
    lambdas.append(p)

for g in range(0,districts):
    for lambder_no in lambdas:
        aggregator(g,lambder_no)

print "\nFirst few tuples with market access stats:"
print population_tuples_list[0:test_tuples]

###### PRINT FINAL ROWS TO CSV #########

headers = ['In_District','Dist_Name','Population']
for l in lambder_list:
    new_column_name = 'MA_lambda_%f' % l
    headers.append(new_column_name)
print headers

with open('output.csv', 'wb') as csvfile:
    output = csv.writer(csvfile, delimiter=',')
    output.writerow(headers)
    for z in range (0,len(population_tuples_list)):
        output.writerow(population_tuples_list[z])
